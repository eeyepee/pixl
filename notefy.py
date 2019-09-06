#! /usr/bin/python
import os
import openpyxl
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

def sendemails():
	smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
	type(smtpObj)
#smtpObj.ehlo()
	smtpObj.starttls()
	smtpObj.login('me@myemail.com', 'mypassword')
	smtpObj.sendmail(me, you , msg.as_string())
	smtpObj.quit()
# me == my email address
# you == recipient's email address
me = "me@gmail.com"
you = "you@gorannet.net"

# Create message container - the correct MIME type is multipart/alternative.
msg = MIMEMultipart('alternative')
msg['Subject'] = "Warehouse shortage"
msg['From'] = me@youremail.com
msg['To'] = you@youremail.com


#----------
wb = openpyxl.load_workbook('workbook.xlsx')
ws = wb["Sheet1"]

slist = []

for i in range(2,ws.max_row+1,1):
    items=ws.cell(row=i, column=1).value
    qty=ws.cell(row=i, column=2).value
    
    # Check qty value type
    if type(qty) is not int:
    	continue
    # Check qty
    if qty < 5:
    	slist.append((items,qty))

print("the list ")
print("-------------------------------------------------------\n")
#print(slist)
print("The below items stock is LOW based on the set threshold:")
print("-------------------------------------------------------\n")
for a,b in slist:
	print(a,b)
#-----------

# Create the body of the message (a plain-text and an HTML version).
text = "Subject: Warehouse Inventory\nHi ,\nRecords show that you have shortage in your warehouse items. Please make this payment as soon as possible. Thank you!'" %(slist)
html = """\
<html>
  <head></head>
  <body>
    <p>Hi!<br><br>
       Records show that you have shortage in your warehouse items. <br><br>

       <b>""" + str(slist) + """</br> <br>
    </p>
  </body>
</html>
"""

# Record the MIME types of both parts - text/plain and text/html.
part1 = MIMEText(text, 'plain')
part2 = MIMEText(html, 'html')

# Attach parts into message container.
# According to RFC 2046, the last part of a multipart message, in this case
# the HTML message, is best and preferred.
msg.attach(part1)
msg.attach(part2)

sendemails()
